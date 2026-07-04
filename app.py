import os
import csv
import io
import psycopg2
import psycopg2.extras
import logging
from logging.handlers import RotatingFileHandler
from datetime import date, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, g, Response
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

# Load env vars
load_dotenv()

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

app = Flask(__name__)

# Configure Logging to File (log errors to app.log)
file_handler = RotatingFileHandler('app.log', maxBytes=10240000, backupCount=5)
file_handler.setLevel(logging.ERROR)
file_formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')
file_handler.setFormatter(file_formatter)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.ERROR)

# Load config from config.py based on environment
env = os.environ.get('FLASK_ENV', 'development')
if env == 'production':
    app.config.from_object('config.ProductionConfig')
else:
    app.config.from_object('config.DevelopmentConfig')

# Configure SQLAlchemy Parameters
app.config['SQLALCHEMY_DATABASE_URI'] = app.config.get('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize SQLAlchemy ORM engine
from flask_sqlalchemy import SQLAlchemy
db_orm = SQLAlchemy(app)

# Mail & SMS configurations
from flask_mail import Mail, Message
from twilio.rest import Client

mail = Mail(app)

UPLOAD_FOLDER = os.path.join('static', 'images')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2MB upload limit

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

class PostgresCursorWrapper:
    def __init__(self, cursor):
        self.cursor = cursor

    def execute(self, sql, params=None):
        # Translate ? to %s for PostgreSQL compatibility
        sql = sql.replace('?', '%s')
        return self.cursor.execute(sql, params)

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()

    def __iter__(self):
        return iter(self.cursor)

    @property
    def rowcount(self):
        return self.cursor.rowcount

    def __getattr__(self, name):
        return getattr(self.cursor, name)

class PostgresConnectionWrapper:
    def __init__(self, conn):
        self.conn = conn

    def cursor(self, *args, **kwargs):
        cursor = self.conn.cursor(*args, **kwargs)
        return PostgresCursorWrapper(cursor)

    def commit(self):
        return self.conn.commit()

    def rollback(self):
        return self.conn.rollback()

    def close(self):
        return self.conn.close()

    def execute(self, sql, params=None):
        cursor = self.cursor()
        cursor.execute(sql, params)
        return cursor

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        conn = psycopg2.connect(app.config['DATABASE_URL'], cursor_factory=psycopg2.extras.DictCursor)
        db = g._database = PostgresConnectionWrapper(conn)
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        db = get_db()
        # Create users table
        db.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(255) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                role VARCHAR(50) NOT NULL CHECK(role IN ('admin', 'staff'))
            )
        ''')
        
        # Create warehouses table
        db.execute('''
            CREATE TABLE IF NOT EXISTS warehouses (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                location VARCHAR(255) NOT NULL,
                manager_name VARCHAR(255) NOT NULL
            )
        ''')

        # Check and create default warehouse
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM warehouses")
        if cursor.fetchone()[0] == 0:
            db.execute(
                "INSERT INTO warehouses (name, location, manager_name) VALUES (?, ?, ?)",
                ('Main Warehouse', 'Central Facility', 'John Doe')
            )
            db.commit()

        # Create products table
        db.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                category VARCHAR(255) NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0,
                price DOUBLE PRECISION NOT NULL DEFAULT 0.0,
                image VARCHAR(255),
                min_stock_threshold INTEGER NOT NULL DEFAULT 5,
                expiry_date DATE,
                warehouse_id INTEGER REFERENCES warehouses(id) DEFAULT 1,
                barcode VARCHAR(255)
            )
        ''')

        # Create stock_movements table
        db.execute('''
            CREATE TABLE IF NOT EXISTS stock_movements (
                id SERIAL PRIMARY KEY,
                product_id INTEGER NOT NULL,
                movement_type VARCHAR(10) NOT NULL CHECK(movement_type IN ('in', 'out')),
                quantity INTEGER NOT NULL,
                note TEXT,
                user_id INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                warehouse_id INTEGER REFERENCES warehouses(id) DEFAULT 1,
                FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        ''')

        # Create suppliers table
        db.execute('''
            CREATE TABLE IF NOT EXISTS suppliers (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                contact_number VARCHAR(100),
                email VARCHAR(255),
                address TEXT
            )
        ''')

        # Create purchase_orders table
        db.execute('''
            CREATE TABLE IF NOT EXISTS purchase_orders (
                id SERIAL PRIMARY KEY,
                supplier_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                order_date VARCHAR(50) NOT NULL,
                payment_status VARCHAR(50) NOT NULL DEFAULT 'pending'
                    CHECK(payment_status IN ('pending','paid')),
                order_status VARCHAR(50) NOT NULL DEFAULT 'pending'
                    CHECK(order_status IN ('pending','received')),
                FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE,
                FOREIGN KEY (product_id)  REFERENCES products(id)  ON DELETE CASCADE
            )
        ''')

        # Create invoices table
        db.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                unit_price DOUBLE PRECISION NOT NULL,
                total_amount DOUBLE PRECISION NOT NULL,
                customer_name VARCHAR(255) NOT NULL,
                date VARCHAR(50) NOT NULL,
                created_by INTEGER NOT NULL,
                FOREIGN KEY (product_id) REFERENCES products(id),
                FOREIGN KEY (created_by) REFERENCES users(id)
            )
        ''')

        # Create settings table
        db.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                key VARCHAR(255) PRIMARY KEY,
                value TEXT
            )
        ''')
        db.commit()

        # Check and create default admin user
        cursor = db.cursor()
        cursor.execute("SELECT * FROM users WHERE username = ?", ('admin',))
        admin_user = cursor.fetchone()
        if not admin_user:
            hashed_pw = generate_password_hash('admin123')
            db.execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                ('admin', hashed_pw, 'admin')
            )
            db.commit()
            print("Default admin user created: admin / admin123")


# Initialize the database on startup
try:
    init_db()
except Exception as e:
    print(f"Warning: Database initialization deferred (connection issue on startup): {e}")

# Helper to read settings
def get_setting(key, default=''):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
    row = cursor.fetchone()
    return row['value'] if row else default

def send_notification_email(subject, body):
    recipient = get_setting('admin_email')
    if not recipient:
        print(f"Skipping email notification (no recipient email configured). Subject: {subject}")
        return False
    try:
        msg = Message(subject, recipients=[recipient])
        msg.body = body
        mail.send(msg)
        print(f"Notification email successfully sent to {recipient}")
        return True
    except Exception as e:
        print(f"Failed to send notification email to {recipient}: {e}")
        return False

def send_notification_sms(body):
    to_phone = get_setting('admin_phone_number')
    sid = get_setting('twilio_sid')
    token = get_setting('twilio_auth_token')
    from_phone = get_setting('twilio_phone_number')
    
    if not to_phone or not sid or not token or not from_phone:
        print("Skipping Twilio SMS (credentials or recipient phone number not fully configured)")
        return False
    try:
        client = Client(sid, token)
        message = client.messages.create(
            body=body,
            from_=from_phone,
            to=to_phone
        )
        print(f"SMS notification successfully sent to {to_phone}: {message.sid}")
        return True
    except Exception as e:
        print(f"Failed to send Twilio SMS notification: {e}")
        return False


# Custom login_required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin-only decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/health')
def health():
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchone()
        return {
            "status": "healthy",
            "database": "connected"
        }, 200
    except Exception as e:
        app.logger.error("Health check failed - Database is disconnected: %s", str(e), exc_info=e)
        return {
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e)
        }, 500

@app.errorhandler(Exception)
def handle_exception(e):
    app.logger.error("Unhandled Exception: %s", str(e), exc_info=e)
    if app.debug:
        raise e
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    return "Internal Server Error", 500

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    # Self-registration always creates a staff account.
    # Admins can create accounts with any role via /users/add.
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash('Username and password are required.', 'danger')
            return render_template('register.html')

        db = get_db()
        try:
            hashed_pw = generate_password_hash(password)
            db.execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                (username, hashed_pw, 'staff')  # always staff on self-registration
            )
            db.commit()
            flash('Account created! You have been registered as Staff. Log in to continue.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username already exists. Please choose a different one.', 'danger')

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash('Username and password are required.', 'danger')
            return render_template('login.html')

        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = cursor.fetchone()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash(f'Welcome back, {user["username"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'danger')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()
    cursor = db.cursor()
    
    # 1. Total Products
    cursor.execute("SELECT COUNT(*) FROM products")
    total_products = cursor.fetchone()[0]
    
    # 2. Low Stock Items — count + full list for the alert panel
    cursor.execute("""
        SELECT id, name, category, quantity, min_stock_threshold
        FROM products
        WHERE quantity < min_stock_threshold
        ORDER BY (min_stock_threshold - quantity) DESC
    """)
    low_stock_products = cursor.fetchall()
    low_stock_items = len(low_stock_products)

    # 3. Stock In Today (sum of all 'in' movements today)
    cursor.execute("""
        SELECT COALESCE(SUM(quantity), 0) FROM stock_movements
        WHERE movement_type = 'in' AND created_at::date = CURRENT_DATE
    """)
    stock_in_today = cursor.fetchone()[0]

    # 4. Stock Out Today (sum of all 'out' movements today)
    cursor.execute("""
        SELECT COALESCE(SUM(quantity), 0) FROM stock_movements
        WHERE movement_type = 'out' AND created_at::date = CURRENT_DATE
    """)
    stock_out_today = cursor.fetchone()[0]

    # 5. Recent movements for activity log (last 10)
    cursor.execute("""
        SELECT sm.*, p.name AS product_name, u.username
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        JOIN users u ON sm.user_id = u.id
        ORDER BY sm.created_at DESC
        LIMIT 10
    """)
    recent_movements = cursor.fetchall()

    # 6. Expiry alerts — expired products
    cursor.execute("""
        SELECT id, name, category, quantity, expiry_date
        FROM products
        WHERE expiry_date IS NOT NULL
          AND expiry_date < CURRENT_DATE
        ORDER BY expiry_date ASC
    """)
    expired_products = cursor.fetchall()

    # 7. Expiring soon — expires within 30 days (but not yet expired)
    cursor.execute("""
        SELECT id, name, category, quantity, expiry_date,
               (expiry_date - CURRENT_DATE) AS days_left
        FROM products
        WHERE expiry_date IS NOT NULL
          AND expiry_date >= CURRENT_DATE
          AND expiry_date <= CURRENT_DATE + INTERVAL '30 days'
        ORDER BY expiry_date ASC
    """)
    expiring_soon = cursor.fetchall()

    # 8. Warehouse-wise stock summary
    cursor.execute("""
        SELECT w.id, w.name, w.location, w.manager_name,
               COUNT(p.id) as total_products,
               COALESCE(SUM(p.quantity), 0) as total_quantity,
               COALESCE(SUM(p.quantity * p.price), 0.0) as total_value
        FROM warehouses w
        LEFT JOIN products p ON p.warehouse_id = w.id
        GROUP BY w.id
        ORDER BY w.name
    """)
    warehouse_summary = cursor.fetchall()
    
    # 9. Top 5 best selling products (based on Stock Out transactions)
    cursor.execute("""
        SELECT p.name, SUM(sm.quantity) as total_sold
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        WHERE sm.movement_type = 'out'
        GROUP BY sm.product_id
        ORDER BY total_sold DESC
        LIMIT 5
    """)
    top_selling = cursor.fetchall()
    top_selling_labels = [row['name'] for row in top_selling]
    top_selling_values = [row['total_sold'] for row in top_selling]

    return render_template('dashboard.html',
                           total_products=total_products,
                           low_stock_items=low_stock_items,
                           low_stock_products=low_stock_products,
                           stock_in_today=stock_in_today,
                           stock_out_today=stock_out_today,
                           recent_movements=recent_movements,
                           expired_products=expired_products,
                           expiring_soon=expiring_soon,
                           warehouse_summary=warehouse_summary,
                           top_selling_labels=top_selling_labels,
                           top_selling_values=top_selling_values)

# --- Products Module ---

@app.route('/products')
@login_required
def products():
    db = get_db()
    cursor = db.cursor()
    
    # Date strings for expiry comparison in template
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    soon_str  = (today + timedelta(days=30)).strftime('%Y-%m-%d')

    # Get all categories for filter dropdown
    cursor.execute("SELECT DISTINCT category FROM products ORDER BY category")
    categories = [row['category'] for row in cursor.fetchall()]

    # Get all warehouses for filter dropdown
    cursor.execute("SELECT id, name FROM warehouses ORDER BY name")
    warehouses = cursor.fetchall()
    
    # Search and Filter Parameters
    search_query = request.args.get('q', '').strip()
    category_filter = request.args.get('category', '').strip()
    status_filter = request.args.get('status', '').strip()
    warehouse_filter = request.args.get('warehouse_id', '').strip()
    
    # Build query dynamically
    query = """
        SELECT p.*, w.name AS warehouse_name
        FROM products p
        JOIN warehouses w ON p.warehouse_id = w.id
        WHERE 1=1
    """
    params = []
    
    if search_query:
        query += " AND p.name LIKE ?"
        params.append(f"%{search_query}%")
        
    if category_filter:
        query += " AND p.category = ?"
        params.append(category_filter)
        
    if status_filter == 'low_stock':
        query += " AND p.quantity < p.min_stock_threshold"
    elif status_filter == 'expiring_soon':
        query += " AND p.expiry_date IS NOT NULL"\
                 " AND p.expiry_date >= CURRENT_DATE"\
                 " AND p.expiry_date <= CURRENT_DATE + INTERVAL '30 days'"
    elif status_filter == 'expired':
        query += " AND p.expiry_date IS NOT NULL"\
                 " AND p.expiry_date < CURRENT_DATE"

    if warehouse_filter:
        query += " AND p.warehouse_id = ?"
        params.append(warehouse_filter)
        
    query += " ORDER BY p.name"
    
    cursor.execute(query, params)
    product_list = cursor.fetchall()
    
    return render_template('products.html', 
                           products=product_list, 
                           categories=categories,
                           warehouses=warehouses,
                           search_query=search_query,
                           category_filter=category_filter,
                           status_filter=status_filter,
                           warehouse_filter=warehouse_filter,
                           today_str=today_str,
                           soon_str=soon_str)

@app.route('/products/add', methods=['POST'])
@admin_required
def add_product():
    db = get_db()
    cursor = db.cursor()

    name = request.form.get('name', '').strip()
    category = request.form.get('category', '').strip()
    quantity = int(request.form.get('quantity', 0))
    price = float(request.form.get('price', 0.0))
    min_stock_threshold = int(request.form.get('min_stock_threshold', 5))
    expiry_date = request.form.get('expiry_date', '').strip() or None
    warehouse_id = int(request.form.get('warehouse_id', 1))
    barcode = request.form.get('barcode', '').strip() or None
    
    if not name or not category:
        flash('Product name and category are required.', 'danger')
        return redirect(url_for('products'))

    if barcode:
        cursor.execute("SELECT id FROM products WHERE barcode = ?", (barcode,))
        if cursor.fetchone():
            flash('Barcode must be unique. Another product already has this barcode.', 'danger')
            return redirect(url_for('products'))
        
    # Handle Image Upload
    image_filename = None
    file = request.files.get('image')
    if file and file.filename != '':
        if allowed_file(file.filename):
            filename = secure_filename(file.filename)
            import time
            filename = f"{int(time.time())}_{filename}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_filename = filename
        else:
            flash('Invalid image format. Allowed: png, jpg, jpeg, gif, webp.', 'danger')
            return redirect(url_for('products'))
            
    db.execute('''
        INSERT INTO products (name, category, quantity, price, image, min_stock_threshold, expiry_date, warehouse_id, barcode)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (name, category, quantity, price, image_filename, min_stock_threshold, expiry_date, warehouse_id, barcode))
    db.commit()
    
    flash('Product added successfully!', 'success')
    return redirect(url_for('products'))

@app.route('/products/edit/<int:product_id>', methods=['POST'])
@admin_required
def edit_product(product_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    product = cursor.fetchone()
    
    if not product:
        flash('Product not found.', 'danger')
        return redirect(url_for('products'))
        
    name = request.form.get('name', '').strip()
    category = request.form.get('category', '').strip()
    quantity = int(request.form.get('quantity', 0))
    price = float(request.form.get('price', 0.0))
    min_stock_threshold = int(request.form.get('min_stock_threshold', 5))
    expiry_date = request.form.get('expiry_date', '').strip() or None
    warehouse_id = int(request.form.get('warehouse_id', 1))
    barcode = request.form.get('barcode', '').strip() or None
    
    if not name or not category:
        flash('Product name and category are required.', 'danger')
        return redirect(url_for('products'))

    if barcode:
        cursor.execute("SELECT id FROM products WHERE barcode = ? AND id != ?", (barcode, product_id))
        if cursor.fetchone():
            flash('Barcode must be unique. Another product already has this barcode.', 'danger')
            return redirect(url_for('products'))
        
    image_filename = product['image']
    
    # Handle Image Upload
    file = request.files.get('image')
    if file and file.filename != '':
        if allowed_file(file.filename):
            # Delete old image if it exists and is not the default
            if product['image']:
                old_path = os.path.join(app.config['UPLOAD_FOLDER'], product['image'])
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except OSError:
                        pass
                        
            filename = secure_filename(file.filename)
            import time
            filename = f"{int(time.time())}_{filename}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_filename = filename
        else:
            flash('Invalid image format. Allowed: png, jpg, jpeg, gif, webp.', 'danger')
            return redirect(url_for('products'))
            
    db.execute('''
        UPDATE products 
        SET name = ?, category = ?, quantity = ?, price = ?, image = ?, min_stock_threshold = ?,
            expiry_date = ?, warehouse_id = ?, barcode = ?
        WHERE id = ?
    ''', (name, category, quantity, price, image_filename, min_stock_threshold, expiry_date, warehouse_id, barcode, product_id))
    db.commit()
    
    flash('Product updated successfully!', 'success')
    return redirect(url_for('products'))

@app.route('/products/delete/<int:product_id>', methods=['POST'])
@admin_required
def delete_product(product_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    product = cursor.fetchone()
    
    if product:
        # Delete image from disk
        if product['image']:
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], product['image'])
            if os.path.exists(image_path):
                try:
                    os.remove(image_path)
                except OSError:
                    pass
                    
        db.execute("DELETE FROM products WHERE id = ?", (product_id,))
        db.commit()
        flash('Product deleted successfully.', 'success')
    else:
        flash('Product not found.', 'danger')
        
    return redirect(url_for('products'))

# --- Stock Movements Module ---

@app.route('/stock')
@login_required
def stock_movements():
    db = get_db()
    cursor = db.cursor()

    # Fetch all products for the filter dropdown
    cursor.execute("SELECT id, name FROM products ORDER BY name")
    product_list = cursor.fetchall()

    # Filters
    product_filter = request.args.get('product_id', '').strip()
    type_filter    = request.args.get('movement_type', '').strip()

    query = """
        SELECT sm.*, p.name AS product_name, u.username
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        JOIN users u ON sm.user_id = u.id
        WHERE 1=1
    """
    params = []

    if product_filter:
        query += " AND sm.product_id = ?"
        params.append(product_filter)

    if type_filter in ('in', 'out'):
        query += " AND sm.movement_type = ?"
        params.append(type_filter)

    query += " ORDER BY sm.created_at DESC"
    cursor.execute(query, params)
    movements = cursor.fetchall()

    return render_template('stock_movements.html',
                           movements=movements,
                           product_list=product_list,
                           product_filter=product_filter,
                           type_filter=type_filter)


@app.route('/stock/in', methods=['GET', 'POST'])
@login_required
def stock_in():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, name, quantity, barcode FROM products ORDER BY name")
    product_list = cursor.fetchall()

    if request.method == 'POST':
        product_id = int(request.form.get('product_id', 0))
        qty        = int(request.form.get('quantity', 0))
        note       = request.form.get('note', '').strip()

        if not product_id or qty <= 0:
            flash('Please select a product and enter a valid quantity.', 'danger')
            return render_template('stock_in.html', product_list=product_list)

        # Record movement
        db.execute("""
            INSERT INTO stock_movements (product_id, movement_type, quantity, note, user_id)
            VALUES (?, 'in', ?, ?, ?)
        """, (product_id, qty, note, session['user_id']))

        # Update product quantity
        db.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?",
                   (qty, product_id))
        db.commit()

        flash(f'Stock In recorded: +{qty} units.', 'success')
        return redirect(url_for('stock_movements'))

    return render_template('stock_in.html', product_list=product_list)


@app.route('/stock/out', methods=['GET', 'POST'])
@login_required
def stock_out():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, name, quantity, price, barcode FROM products ORDER BY name")
    product_list = cursor.fetchall()

    if request.method == 'POST':
        product_id    = int(request.form.get('product_id', 0))
        qty           = int(request.form.get('quantity', 0))
        note          = request.form.get('note', '').strip()
        customer_name = request.form.get('customer_name', '').strip()

        if not product_id or qty <= 0:
            flash('Please select a product and enter a valid quantity.', 'danger')
            return render_template('stock_out.html', product_list=product_list)

        if not customer_name:
            flash('Customer name is required to generate an invoice.', 'danger')
            return render_template('stock_out.html', product_list=product_list)

        # Check available stock
        cursor.execute("SELECT quantity, price FROM products WHERE id = ?", (product_id,))
        row = cursor.fetchone()
        if not row or row['quantity'] < qty:
            flash('Insufficient stock available for this product.', 'danger')
            return render_template('stock_out.html', product_list=product_list)

        unit_price   = row['price']
        total_amount = round(unit_price * qty, 2)
        today_str    = date.today().strftime('%Y-%m-%d')

        # Record stock movement
        db.execute("""
            INSERT INTO stock_movements (product_id, movement_type, quantity, note, user_id)
            VALUES (?, 'out', ?, ?, ?)
        """, (product_id, qty, note, session['user_id']))

        # Update product quantity
        db.execute("UPDATE products SET quantity = quantity - ? WHERE id = ?",
                   (qty, product_id))

        # Auto-create invoice
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO invoices (product_id, quantity, unit_price, total_amount,
                                  customer_name, date, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?) RETURNING id
        """, (product_id, qty, unit_price, total_amount,
               customer_name, today_str, session['user_id']))
        invoice_id = cursor.fetchone()[0]
        db.commit()

        # Check stock alert levels for notifications
        cursor.execute("SELECT name, quantity, min_stock_threshold FROM products WHERE id = ?", (product_id,))
        p_info = cursor.fetchone()
        if p_info and p_info['quantity'] < p_info['min_stock_threshold']:
            subject = f"⚠️ Low Stock Alert: {p_info['name']}"
            body = (f"System Alert: Product stock level has fallen below threshold.\n\n"
                    f"Product: {p_info['name']}\n"
                    f"Current Stock: {p_info['quantity']}\n"
                    f"Minimum Threshold: {p_info['min_stock_threshold']}\n\n"
                    f"Please contact suppliers to submit a restock Purchase Order.")
            send_notification_email(subject, body)
            
            if p_info['quantity'] <= max(1, p_info['min_stock_threshold'] // 2):
                sms_body = f"[CRITICAL LOW STOCK] Product '{p_info['name']}' is at {p_info['quantity']} units (Threshold: {p_info['min_stock_threshold']}). Restock immediately!"
                send_notification_sms(sms_body)

        flash(f'Stock Out recorded — Invoice #{invoice_id} generated for {customer_name}.', 'success')
        return redirect(url_for('invoices'))

    return render_template('stock_out.html', product_list=product_list)


# --- Invoices ---

@app.route('/invoices')
@login_required
def invoices():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT inv.*, p.name AS product_name, p.category, u.username AS created_by_name
        FROM invoices inv
        JOIN products p ON inv.product_id = p.id
        JOIN users u ON inv.created_by = u.id
        ORDER BY inv.date DESC, inv.id DESC
    """)
    invoice_list = cursor.fetchall()

    # Summary totals
    cursor.execute("SELECT COALESCE(SUM(total_amount), 0) FROM invoices")
    grand_total = cursor.fetchone()[0]

    return render_template('invoices.html',
                           invoices=invoice_list,
                           grand_total=grand_total)


@app.route('/invoices/pdf/<int:invoice_id>')
@login_required
def download_invoice_pdf(invoice_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT inv.*, p.name AS product_name, p.category, p.price,
               u.username AS created_by_name
        FROM invoices inv
        JOIN products p ON inv.product_id = p.id
        JOIN users u ON inv.created_by = u.id
        WHERE inv.id = ?
    """, (invoice_id,))
    inv = cursor.fetchone()

    if not inv:
        flash('Invoice not found.', 'danger')
        return redirect(url_for('invoices'))

    # --- Build PDF in memory ---
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    INDIGO = colors.HexColor('#4f46e5')
    LIGHT  = colors.HexColor('#f4f6f9')
    DARK   = colors.HexColor('#1e1e2e')

    # Custom paragraph styles
    title_style = ParagraphStyle('title', parent=styles['Normal'],
                                 fontSize=26, textColor=INDIGO,
                                 fontName='Helvetica-Bold', leading=32)
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'],
                                    fontSize=10, textColor=colors.grey,
                                    fontName='Helvetica')
    label_style = ParagraphStyle('label', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.grey,
                                 fontName='Helvetica')
    value_style = ParagraphStyle('value', parent=styles['Normal'],
                                 fontSize=11, textColor=DARK,
                                 fontName='Helvetica-Bold')
    footer_style = ParagraphStyle('footer', parent=styles['Normal'],
                                  fontSize=8, textColor=colors.grey,
                                  alignment=TA_CENTER)

    story = []

    # ── Header band ──
    header_data = [[
        Paragraph('<b>Stock Manager</b>', title_style),
        Paragraph(f'INVOICE <font color="#4f46e5">#{inv["id"]:04d}</font>',
                  ParagraphStyle('inv_num', parent=styles['Normal'],
                                 fontSize=18, fontName='Helvetica-Bold',
                                 textColor=DARK, alignment=TA_RIGHT))
    ]]
    header_table = Table(header_data, colWidths=[10*cm, 7*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(header_table)
    story.append(Paragraph('Automated Invoice System', subtitle_style))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width='100%', thickness=2, color=INDIGO, spaceAfter=0.4*cm))

    # ── Bill-to / Invoice meta ──
    meta_data = [
        [Paragraph('BILLED TO', label_style),   Paragraph('', label_style),
         Paragraph('INVOICE DATE', label_style), Paragraph('', label_style)],
        [Paragraph(inv['customer_name'], value_style), Paragraph('', label_style),
         Paragraph(inv['date'], value_style),    Paragraph('', label_style)],
        [Paragraph('', label_style), Paragraph('', label_style),
         Paragraph('ISSUED BY', label_style),   Paragraph('', label_style)],
        [Paragraph('', label_style), Paragraph('', label_style),
         Paragraph(inv['created_by_name'], value_style), Paragraph('', label_style)],
    ]
    meta_table = Table(meta_data, colWidths=[8*cm, 1*cm, 5*cm, 3*cm])
    meta_table.setStyle(TableStyle([
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Line items table ──
    items_header = ['#', 'Product', 'Category', 'Unit Price', 'Qty', 'Total']
    items_data   = [items_header, [
        '1',
        inv['product_name'],
        inv['category'],
        f"${inv['unit_price']:.2f}",
        str(inv['quantity']),
        f"${inv['total_amount']:.2f}",
    ]]
    items_table = Table(items_data, colWidths=[1*cm, 5.5*cm, 3*cm, 2.5*cm, 1.5*cm, 3.5*cm])
    items_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND',    (0,0), (-1,0), INDIGO),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 9),
        ('ALIGN',         (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING',    (0,0), (-1,0), 8),
        # Data rows
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,1), (-1,-1), 10),
        ('ALIGN',         (3,1), (-1,-1), 'RIGHT'),
        ('ALIGN',         (0,1), (2,-1), 'LEFT'),
        ('BACKGROUND',    (0,1), (-1,-1), LIGHT),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [LIGHT, colors.white]),
        ('TOPPADDING',    (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 0.4*cm))

    # ── Totals section ──
    totals_data = [
        ['', 'Subtotal:', f"${inv['total_amount']:.2f}"],
        ['', 'Tax (0%):', '$0.00'],
        ['', Paragraph('<b>TOTAL DUE</b>', ParagraphStyle('t', parent=styles['Normal'],
                        fontSize=13, fontName='Helvetica-Bold', textColor=INDIGO)),
              Paragraph(f'<b>${inv["total_amount"]:.2f}</b>',
                        ParagraphStyle('tv', parent=styles['Normal'],
                                       fontSize=13, fontName='Helvetica-Bold',
                                       textColor=INDIGO, alignment=TA_RIGHT))],
    ]
    totals_table = Table(totals_data, colWidths=[9.5*cm, 4*cm, 3.5*cm])
    totals_table.setStyle(TableStyle([
        ('ALIGN',         (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME',      (1,0), (1,1), 'Helvetica'),
        ('FONTNAME',      (2,0), (2,1), 'Helvetica'),
        ('FONTSIZE',      (0,0), (-1,-1), 10),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEABOVE',     (1,2), (-1,2), 1.5, INDIGO),
        ('TOPPADDING',    (0,2), (-1,2), 8),
    ]))
    story.append(totals_table)

    # ── Footer ──
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.lightgrey, spaceAfter=0.3*cm))
    story.append(Paragraph(
        'Thank you for your business!  •  This invoice was auto-generated by Stock Manager.',
        footer_style
    ))

    doc.build(story)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename=invoice_{invoice_id:04d}.pdf'
        }
    )


# --- Reports & Analytics ---

@app.route('/reports')
@admin_required
def reports():
    db = get_db()
    cursor = db.cursor()

    # Get filter params
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    # Base conditions for filtering stock movements
    movements_cond = "1=1"
    params = []
    if start_date:
        movements_cond += " AND sm.created_at::date >= ?"
        params.append(start_date)
    if end_date:
        movements_cond += " AND sm.created_at::date <= ?"
        params.append(end_date)

    # 1. Bar chart: top 5 best-selling products (based on Stock Out transactions)
    best_sellers_query = f"""
        SELECT p.name, SUM(sm.quantity) as total_sold
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        WHERE sm.movement_type = 'out' AND {movements_cond}
        GROUP BY sm.product_id, p.name
        ORDER BY total_sold DESC
        LIMIT 5
    """
    cursor.execute(best_sellers_query, params)
    best_sellers_rows = cursor.fetchall()
    best_sellers_labels = [r['name'] for r in best_sellers_rows]
    best_sellers_data = [r['total_sold'] for r in best_sellers_rows]

    # 2. Line chart: monthly stock movement (total stock in vs stock out per month)
    monthly_query = f"""
        SELECT to_char(sm.created_at, 'YYYY-MM') as month,
               SUM(CASE WHEN sm.movement_type='in'  THEN sm.quantity ELSE 0 END) as total_in,
               SUM(CASE WHEN sm.movement_type='out' THEN sm.quantity ELSE 0 END) as total_out
        FROM stock_movements sm
        WHERE {movements_cond}
        GROUP BY to_char(sm.created_at, 'YYYY-MM')
        ORDER BY to_char(sm.created_at, 'YYYY-MM')
    """
    cursor.execute(monthly_query, params)
    monthly_rows = cursor.fetchall()
    monthly_labels = [r['month'] for r in monthly_rows]
    monthly_in_data = [r['total_in'] for r in monthly_rows]
    monthly_out_data = [r['total_out'] for r in monthly_rows]

    # 3. Stock value by category
    cursor.execute("""
        SELECT category,
               COUNT(*) as product_count,
               SUM(quantity * price) as total_value
        FROM products
        GROUP BY category
        ORDER BY total_value DESC
    """)
    cat_rows = cursor.fetchall()
    cat_labels = [r['category']    for r in cat_rows]
    cat_values = [round(r['total_value'] or 0, 2) for r in cat_rows]
    cat_counts = [r['product_count'] for r in cat_rows]

    # 4. Summary totals
    cursor.execute("SELECT COUNT(*), SUM(quantity * price) FROM products")
    row = cursor.fetchone()
    total_products  = row[0] or 0
    total_inventory_value = round(row[1] or 0, 2)

    # Count transaction types within the filtered range
    cursor.execute(f"SELECT COUNT(*) FROM stock_movements sm WHERE sm.movement_type='in' AND {movements_cond}", params)
    total_in_txns = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM stock_movements sm WHERE sm.movement_type='out' AND {movements_cond}", params)
    total_out_txns = cursor.fetchone()[0]

    return render_template('reports.html',
                           start_date=start_date,
                           end_date=end_date,
                           best_sellers_labels=best_sellers_labels,
                           best_sellers_data=best_sellers_data,
                           monthly_labels=monthly_labels,
                           monthly_in_data=monthly_in_data,
                           monthly_out_data=monthly_out_data,
                           cat_labels=cat_labels,
                           cat_values=cat_values,
                           cat_counts=cat_counts,
                           total_products=total_products,
                           total_inventory_value=total_inventory_value,
                           total_in_txns=total_in_txns,
                           total_out_txns=total_out_txns)


# --- CSV Export ---

@app.route('/export/products')
@admin_required
def export_products():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, name, category, quantity, price, min_stock_threshold FROM products ORDER BY name")
    rows = cursor.fetchall()

    def generate():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['ID', 'Name', 'Category', 'Quantity', 'Price', 'Min Stock Threshold'])
        for r in rows:
            writer.writerow([r['id'], r['name'], r['category'], r['quantity'],
                             f"{r['price']:.2f}", r['min_stock_threshold']])
        yield buf.getvalue()

    return Response(generate(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=products.csv'})


@app.route('/export/transactions')
@admin_required
def export_transactions():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT sm.id, p.name AS product_name, sm.movement_type, sm.quantity,
               sm.note, u.username, sm.created_at
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        JOIN users u ON sm.user_id = u.id
        ORDER BY sm.created_at DESC
    """)
    rows = cursor.fetchall()

    def generate():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['ID', 'Product', 'Type', 'Quantity', 'Note', 'Recorded By', 'Date'])
        for r in rows:
            writer.writerow([r['id'], r['product_name'], r['movement_type'],
                             r['quantity'], r['note'] or '', r['username'], r['created_at']])
        yield buf.getvalue()

    return Response(generate(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=transactions.csv'})


@app.route('/export/products/excel')
@login_required
def export_products_excel():
    db = get_db()
    cursor = db.cursor()
    
    # Get parameters
    search_query = request.args.get('q', '').strip()
    category_filter = request.args.get('category', '').strip()
    status_filter = request.args.get('status', '').strip()
    warehouse_filter = request.args.get('warehouse_id', '').strip()
    
    # Build query dynamically
    query = """
        SELECT p.*, w.name AS warehouse_name
        FROM products p
        JOIN warehouses w ON p.warehouse_id = w.id
        WHERE 1=1
    """
    params = []
    
    if search_query:
        query += " AND p.name LIKE ?"
        params.append(f"%{search_query}%")
        
    if category_filter:
        query += " AND p.category = ?"
        params.append(category_filter)
        
    if status_filter == 'low_stock':
        query += " AND p.quantity < p.min_stock_threshold"
    elif status_filter == 'expiring_soon':
        query += " AND p.expiry_date IS NOT NULL AND p.expiry_date != ''"\
                 " AND DATE(p.expiry_date) >= DATE('now','localtime')"\
                 " AND DATE(p.expiry_date) <= DATE('now','localtime','+30 days')"
    elif status_filter == 'expired':
        query += " AND p.expiry_date IS NOT NULL AND p.expiry_date != ''"\
                 " AND DATE(p.expiry_date) < DATE('now','localtime')"

    if warehouse_filter:
        query += " AND p.warehouse_id = ?"
        params.append(warehouse_filter)
        
    query += " ORDER BY p.name"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Inventory"
    ws.views.sheetView[0].showGridLines = True
    
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Segoe UI', size=10)
    total_font = Font(name='Segoe UI', size=11, bold=True)
    
    indigo_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_fill = PatternFill(start_color='312E81', end_color='312E81', fill_type='solid')
    zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Title
    ws.merge_cells('A1:H2')
    title_cell = ws['A1']
    title_cell.value = "Products Inventory Report"
    title_cell.font = title_font
    title_cell.fill = indigo_fill
    title_cell.alignment = align_center
    
    # Info
    ws['A3'] = "Export Date:"
    ws['A3'].font = Font(name='Segoe UI', size=10, bold=True)
    ws['B3'] = date.today().strftime('%Y-%m-%d')
    ws['B3'].font = data_font
    
    if search_query or category_filter or status_filter or warehouse_filter:
        ws['D3'] = "Active Filters:"
        ws['D3'].font = Font(name='Segoe UI', size=10, bold=True)
        filters = []
        if search_query: filters.append(f"Search: '{search_query}'")
        if category_filter: filters.append(f"Category: {category_filter}")
        if status_filter: filters.append(f"Status: {status_filter}")
        if warehouse_filter:
            cursor.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_filter,))
            wh_row = cursor.fetchone()
            if wh_row:
                filters.append(f"Warehouse: {wh_row['name']}")
        ws['E3'] = ", ".join(filters)
        ws['E3'].font = data_font
        
    headers = ['ID', 'Product Name', 'Category', 'Quantity', 'Price', 'Min Threshold', 'Expiry Date', 'Warehouse']
    ws.append([])
    ws.append(headers)
    
    for col_num in range(1, 9):
        cell = ws.cell(row=5, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    start_row = 6
    total_qty = 0
    total_val = 0.0
    for idx, r in enumerate(rows):
        current_row = start_row + idx
        qty = r['quantity']
        price = r['price']
        total_qty += qty
        total_val += (qty * price)
        
        ws.append([
            r['id'],
            r['name'],
            r['category'],
            qty,
            price,
            r['min_stock_threshold'],
            r['expiry_date'] or '—',
            r['warehouse_name']
        ])
        
        for col_num in range(1, 9):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = zebra_fill
            if col_num in (1, 6, 7):
                cell.alignment = align_center
            elif col_num in (2, 3, 8):
                cell.alignment = align_left
            elif col_num == 4:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_num == 5:
                cell.alignment = align_right
                cell.number_format = '$#,##0.00'
                
    tot_row = start_row + len(rows)
    ws.append([])
    ws.merge_cells(start_row=tot_row + 1, start_column=1, end_row=tot_row + 1, end_column=3)
    ws.cell(row=tot_row + 1, column=1).value = "TOTAL INVENTORY SUMMARY"
    ws.cell(row=tot_row + 1, column=1).font = total_font
    ws.cell(row=tot_row + 1, column=1).alignment = Alignment(horizontal='right', vertical='center')
    
    cell_qty = ws.cell(row=tot_row + 1, column=4)
    cell_qty.value = total_qty
    cell_qty.font = total_font
    cell_qty.alignment = align_right
    cell_qty.number_format = '#,##0'
    cell_qty.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    
    ws.cell(row=tot_row + 1, column=5).value = f"Total Value: ${total_val:,.2f}"
    ws.cell(row=tot_row + 1, column=5).font = total_font
    ws.cell(row=tot_row + 1, column=5).alignment = align_right
    ws.merge_cells(start_row=tot_row + 1, start_column=5, end_row=tot_row + 1, end_column=8)
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2) or cell.row == (tot_row + 1):
                continue
            val_str = str(cell.value or '')
            if cell.number_format == '$#,##0.00' and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.2f}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename=products_inventory.xlsx'
        }
    )


@app.route('/export/reports/pdf')
@admin_required
def export_reports_pdf():
    db = get_db()
    cursor = db.cursor()

    # Get filter params
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    # Base conditions for filtering stock movements
    movements_cond = "1=1"
    params = []
    if start_date:
        movements_cond += " AND sm.created_at::date >= ?"
        params.append(start_date)
    if end_date:
        movements_cond += " AND sm.created_at::date <= ?"
        params.append(end_date)

    # 1. Top 5 best-selling products
    best_sellers_query = f"""
        SELECT p.name, SUM(sm.quantity) as total_sold
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        WHERE sm.movement_type = 'out' AND {movements_cond}
        GROUP BY sm.product_id, p.name
        ORDER BY total_sold DESC
        LIMIT 5
    """
    cursor.execute(best_sellers_query, params)
    best_sellers = cursor.fetchall()

    # 2. Monthly stock movements
    monthly_query = f"""
        SELECT to_char(sm.created_at, 'YYYY-MM') as month,
               SUM(CASE WHEN sm.movement_type='in'  THEN sm.quantity ELSE 0 END) as total_in,
               SUM(CASE WHEN sm.movement_type='out' THEN sm.quantity ELSE 0 END) as total_out
        FROM stock_movements sm
        WHERE {movements_cond}
        GROUP BY to_char(sm.created_at, 'YYYY-MM')
        ORDER BY to_char(sm.created_at, 'YYYY-MM')
    """
    cursor.execute(monthly_query, params)
    monthly_rows = cursor.fetchall()

    # 3. Category breakdown
    cursor.execute("""
        SELECT category,
               COUNT(*) as product_count,
               SUM(quantity * price) as total_value
        FROM products
        GROUP BY category
        ORDER BY total_value DESC
    """)
    categories = cursor.fetchall()

    # 4. KPI Summary totals
    cursor.execute("SELECT COUNT(*), SUM(quantity * price) FROM products")
    row = cursor.fetchone()
    total_products = row[0] or 0
    total_inventory_value = round(row[1] or 0, 2)

    cursor.execute(f"SELECT COUNT(*) FROM stock_movements sm WHERE sm.movement_type='in' AND {movements_cond}", params)
    total_in_txns = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM stock_movements sm WHERE sm.movement_type='out' AND {movements_cond}", params)
    total_out_txns = cursor.fetchone()[0]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    INDIGO = colors.HexColor('#4f46e5')
    DARK   = colors.HexColor('#1f2937')
    LIGHT  = colors.HexColor('#f9fafb')
    BORDER_COLOR = colors.HexColor('#e5e7eb')

    title_style = ParagraphStyle('ReportTitle', parent=styles['Normal'],
                                 fontSize=24, textColor=INDIGO,
                                 fontName='Helvetica-Bold', leading=28)
    subtitle_style = ParagraphStyle('ReportSubtitle', parent=styles['Normal'],
                                    fontSize=10, textColor=colors.HexColor('#6b7280'),
                                    fontName='Helvetica')
    h2_style = ParagraphStyle('SectionHeading', parent=styles['Normal'],
                               fontSize=14, textColor=DARK,
                               fontName='Helvetica-Bold', leading=18,
                               spaceBefore=15, spaceAfter=8)
    label_style = ParagraphStyle('KpiLabel', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.HexColor('#6b7280'),
                                 alignment=TA_CENTER)
    value_style = ParagraphStyle('KpiValue', parent=styles['Normal'],
                                 fontSize=14, textColor=INDIGO,
                                 fontName='Helvetica-Bold', alignment=TA_CENTER)
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'],
                                         fontSize=9, textColor=colors.white,
                                         fontName='Helvetica-Bold')
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'],
                                 fontSize=9, textColor=DARK)
    cell_bold_style = ParagraphStyle('TableCellBold', parent=styles['Normal'],
                                      fontSize=9, textColor=DARK, fontName='Helvetica-Bold')

    story = []

    header_data = [
        [Paragraph('<b>Stock Manager</b>', title_style),
         Paragraph('EXECUTIVE REPORT', ParagraphStyle('RightSub', parent=styles['Normal'],
                                                     fontSize=14, fontName='Helvetica-Bold',
                                                     textColor=DARK, alignment=TA_RIGHT))]
    ]
    header_table = Table(header_data, colWidths=[10*cm, 7*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(header_table)
    
    date_str = date.today().strftime('%B %d, %Y')
    filter_desc = "All Time"
    if start_date and end_date:
        filter_desc = f"{start_date} to {end_date}"
    elif start_date:
        filter_desc = f"From {start_date}"
    elif end_date:
        filter_desc = f"Until {end_date}"
        
    story.append(Paragraph(f'Generated on {date_str}  •  Filters: {filter_desc}', subtitle_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width='100%', thickness=2, color=INDIGO, spaceAfter=0.6*cm))

    kpi_data = [
        [
            Paragraph('Total Products', label_style),
            Paragraph('Inventory Value', label_style),
            Paragraph('Stock-In txns (Filtered)', label_style),
            Paragraph('Stock-Out txns (Filtered)', label_style)
        ],
        [
            Paragraph(f'{total_products}', value_style),
            Paragraph(f'${total_inventory_value:,.2f}', value_style),
            Paragraph(f'{total_in_txns}', value_style),
            Paragraph(f'{total_out_txns}', value_style)
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[4.25*cm, 4.25*cm, 4.25*cm, 4.25*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), LIGHT),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 2),
        ('TOPPADDING', (0,1), (-1,1), 2),
        ('BOTTOMPADDING', (0,1), (-1,1), 8),
        ('BOX', (0,0), (-1,-1), 1, BORDER_COLOR),
        ('INNERGRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.6*cm))

    story.append(Paragraph('Top Best-Selling Products', h2_style))
    if best_sellers:
        bs_header = [Paragraph('Product Name', table_header_style), Paragraph('Total Units Sold', table_header_style)]
        bs_rows = [bs_header]
        for idx, bs in enumerate(best_sellers):
            bs_rows.append([
                Paragraph(bs['name'], cell_style),
                Paragraph(f"{bs['total_sold']}", cell_bold_style)
            ])
        bs_table = Table(bs_rows, colWidths=[11*cm, 6*cm])
        bs_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), INDIGO),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [LIGHT, colors.white]),
            ('BOX', (0,0), (-1,-1), 1, BORDER_COLOR),
            ('INNERGRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ]))
        story.append(bs_table)
    else:
        story.append(Paragraph('No best-seller sales data found for the selected period.', subtitle_style))
    
    story.append(Spacer(1, 0.6*cm))

    story.append(Paragraph('Monthly Stock Movement', h2_style))
    if monthly_rows:
        m_header = [Paragraph('Month', table_header_style), Paragraph('Stock In (Qty)', table_header_style), Paragraph('Stock Out (Qty)', table_header_style)]
        m_data = [m_header]
        for m in monthly_rows:
            m_data.append([
                Paragraph(m['month'], cell_style),
                Paragraph(f"+{m['total_in']}", cell_style),
                Paragraph(f"-{m['total_out']}", cell_style)
            ])
        m_table = Table(m_data, colWidths=[6*cm, 5.5*cm, 5.5*cm])
        m_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), INDIGO),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [LIGHT, colors.white]),
            ('BOX', (0,0), (-1,-1), 1, BORDER_COLOR),
            ('INNERGRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(m_table)
    else:
        story.append(Paragraph('No stock movement transactions found for the selected period.', subtitle_style))

    story.append(Spacer(1, 0.6*cm))

    story.append(Paragraph('Category Breakdown', h2_style))
    if categories:
        cat_header = [Paragraph('Category Name', table_header_style), Paragraph('Product Count', table_header_style), Paragraph('Total Value', table_header_style)]
        cat_data = [cat_header]
        for c in categories:
            cat_data.append([
                Paragraph(c['category'], cell_style),
                Paragraph(f"{c['product_count']}", cell_style),
                Paragraph(f"${c['total_value'] or 0:,.2f}", cell_bold_style)
            ])
        cat_table = Table(cat_data, colWidths=[7*cm, 4*cm, 6*cm])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), INDIGO),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [LIGHT, colors.white]),
            ('BOX', (0,0), (-1,-1), 1, BORDER_COLOR),
            ('INNERGRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(cat_table)
    else:
        story.append(Paragraph('No product category data found.', subtitle_style))

    story.append(Spacer(1, 1.2*cm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.lightgrey, spaceAfter=0.3*cm))
    story.append(Paragraph('Stock Manager System  •  Executive Analytics Summary', ParagraphStyle('Footer', parent=styles['Normal'], alignment=TA_CENTER, fontSize=8, textColor=colors.grey)))

    doc.build(story)
    buf.seek(0)

    filename = f"executive_report_{start_date or 'all'}_to_{end_date or 'all'}.pdf"
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename={filename}'
        }
    )


# --- User Management (admin only) ---

@app.route('/users')
@admin_required
def manage_users():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, username, role FROM users ORDER BY username")
    users = cursor.fetchall()
    return render_template('manage_users.html', users=users)


@app.route('/users/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'staff')

        if not username or not password:
            flash('Username and password are required.', 'danger')
            return render_template('add_user.html')

        if role not in ['admin', 'staff']:
            flash('Invalid role selected.', 'danger')
            return render_template('add_user.html')

        db = get_db()
        try:
            hashed_pw = generate_password_hash(password)
            db.execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                (username, hashed_pw, role)
            )
            db.commit()
            flash('User created successfully!', 'success')
            return redirect(url_for('manage_users'))
        except sqlite3.IntegrityError:
            flash('Username already exists. Please choose a different one.', 'danger')

    return render_template('add_user.html')


@app.route('/users/change-role/<int:user_id>', methods=['POST'])
@admin_required
def change_user_role(user_id):
    if user_id == session['user_id']:
        flash('You cannot change your own role.', 'warning')
        return redirect(url_for('manage_users'))

    new_role = request.form.get('role')
    if new_role not in ('admin', 'staff'):
        flash('Invalid role.', 'danger')
        return redirect(url_for('manage_users'))

    db = get_db()
    db.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
    db.commit()
    flash(f'User role updated to {new_role}.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    if user_id == session['user_id']:
        flash('You cannot delete your own account.', 'warning')
        return redirect(url_for('manage_users'))

    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    flash('User deleted successfully.', 'success')
    return redirect(url_for('manage_users'))


# ================================================================
# --- Suppliers Module ---
# ================================================================

@app.route('/suppliers')
@admin_required
def suppliers():
    db = get_db()
    cursor = db.cursor()
    q = request.args.get('q', '').strip()
    if q:
        cursor.execute(
            "SELECT * FROM suppliers WHERE name LIKE ? OR email LIKE ? ORDER BY name",
            (f'%{q}%', f'%{q}%')
        )
    else:
        cursor.execute("SELECT * FROM suppliers ORDER BY name")
    supplier_list = cursor.fetchall()
    return render_template('suppliers.html', suppliers=supplier_list, q=q)


@app.route('/suppliers/add', methods=['GET', 'POST'])
@admin_required
def add_supplier():
    if request.method == 'POST':
        name           = request.form.get('name', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        email          = request.form.get('email', '').strip()
        address        = request.form.get('address', '').strip()
        if not name:
            flash('Supplier name is required.', 'danger')
            return render_template('add_supplier.html')
        db = get_db()
        db.execute(
            "INSERT INTO suppliers (name, contact_number, email, address) VALUES (?,?,?,?)",
            (name, contact_number, email, address)
        )
        db.commit()
        flash('Supplier added successfully!', 'success')
        return redirect(url_for('suppliers'))
    return render_template('add_supplier.html')


@app.route('/suppliers/edit/<int:supplier_id>', methods=['GET', 'POST'])
@admin_required
def edit_supplier(supplier_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM suppliers WHERE id = ?", (supplier_id,))
    supplier = cursor.fetchone()
    if not supplier:
        flash('Supplier not found.', 'danger')
        return redirect(url_for('suppliers'))
    if request.method == 'POST':
        name           = request.form.get('name', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        email          = request.form.get('email', '').strip()
        address        = request.form.get('address', '').strip()
        if not name:
            flash('Supplier name is required.', 'danger')
            return render_template('edit_supplier.html', supplier=supplier)
        db.execute(
            "UPDATE suppliers SET name=?, contact_number=?, email=?, address=? WHERE id=?",
            (name, contact_number, email, address, supplier_id)
        )
        db.commit()
        flash('Supplier updated successfully!', 'success')
        return redirect(url_for('suppliers'))
    return render_template('edit_supplier.html', supplier=supplier)


@app.route('/suppliers/delete/<int:supplier_id>', methods=['POST'])
@admin_required
def delete_supplier(supplier_id):
    db = get_db()
    db.execute("DELETE FROM suppliers WHERE id = ?", (supplier_id,))
    db.commit()
    flash('Supplier deleted.', 'success')
    return redirect(url_for('suppliers'))


# ================================================================
# --- Purchase Orders Module ---
# ================================================================

@app.route('/purchase-orders')
@admin_required
def purchase_orders():
    db = get_db()
    cursor = db.cursor()

    pay_filter   = request.args.get('payment_status', '').strip()
    order_filter = request.args.get('order_status', '').strip()

    query = """
        SELECT po.*,
               s.name  AS supplier_name,
               p.name  AS product_name
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.id
        JOIN products  p ON po.product_id  = p.id
        WHERE 1=1
    """
    params = []
    if pay_filter in ('pending', 'paid'):
        query += " AND po.payment_status = ?"
        params.append(pay_filter)
    if order_filter in ('pending', 'received'):
        query += " AND po.order_status = ?"
        params.append(order_filter)
    query += " ORDER BY po.order_date DESC, po.id DESC"

    cursor.execute(query, params)
    orders = cursor.fetchall()
    return render_template('purchase_orders.html',
                           orders=orders,
                           pay_filter=pay_filter,
                           order_filter=order_filter)


@app.route('/purchase-orders/add', methods=['GET', 'POST'])
@admin_required
def add_purchase_order():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
    supplier_list = cursor.fetchall()
    cursor.execute("SELECT id, name FROM products ORDER BY name")
    product_list = cursor.fetchall()

    if request.method == 'POST':
        import datetime
        supplier_id = request.form.get('supplier_id', 0)
        product_id  = request.form.get('product_id', 0)
        quantity    = request.form.get('quantity', 0)
        order_date  = request.form.get('order_date', '')

        if not supplier_id or not product_id or not quantity or not order_date:
            flash('All fields are required.', 'danger')
            return render_template('add_purchase_order.html',
                                   supplier_list=supplier_list,
                                   product_list=product_list)
        db.execute("""
            INSERT INTO purchase_orders
                (supplier_id, product_id, quantity, order_date, payment_status, order_status)
            VALUES (?, ?, ?, ?, 'pending', 'pending')
        """, (int(supplier_id), int(product_id), int(quantity), order_date))
        db.commit()
        flash('Purchase order created successfully!', 'success')
        return redirect(url_for('purchase_orders'))

    today = __import__('datetime').date.today().isoformat()
    return render_template('add_purchase_order.html',
                           supplier_list=supplier_list,
                           product_list=product_list,
                           today=today)


@app.route('/purchase-orders/update/<int:order_id>', methods=['POST'])
@admin_required
def update_purchase_order(order_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM purchase_orders WHERE id = ?", (order_id,))
    order = cursor.fetchone()
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('purchase_orders'))

    new_payment = request.form.get('payment_status', order['payment_status'])
    new_order   = request.form.get('order_status',  order['order_status'])

    if new_payment not in ('pending', 'paid'):
        new_payment = order['payment_status']
    if new_order not in ('pending', 'received'):
        new_order = order['order_status']

    # If order just became 'received' (was pending before), add a stock-in movement
    if new_order == 'received' and order['order_status'] == 'pending':
        db.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?",
                   (order['quantity'], order['product_id']))
        
        cursor.execute("SELECT warehouse_id FROM products WHERE id = ?", (order['product_id'],))
        p_row = cursor.fetchone()
        wh_id = p_row['warehouse_id'] if p_row else 1

        db.execute("""
            INSERT INTO stock_movements
                (product_id, movement_type, quantity, note, user_id, warehouse_id)
            VALUES (?, 'in', ?, 'Purchase Order #' || ?, ?, ?)
        """, (order['product_id'], order['quantity'], order_id, session['user_id'], wh_id))

        # Send Email Alert for PO Received
        cursor.execute("SELECT name FROM products WHERE id = ?", (order['product_id'],))
        p_name = cursor.fetchone()[0]
        subject = f"📦 Purchase Order #{order_id} Received"
        body = (f"System Notification: Purchase Order #{order_id} has been marked as RECEIVED.\n\n"
                f"Product: {p_name}\n"
                f"Quantity: {order['quantity']}\n"
                f"Date: {__import__('datetime').date.today().isoformat()}\n\n"
                f"Inventory stock has been automatically updated (+{order['quantity']} units).")
        send_notification_email(subject, body)

    db.execute("""
        UPDATE purchase_orders
        SET payment_status = ?, order_status = ?
        WHERE id = ?
    """, (new_payment, new_order, order_id))
    db.commit()
    flash('Order status updated.', 'success')
    return redirect(url_for('purchase_orders'))


@app.route('/purchase-orders/delete/<int:order_id>', methods=['POST'])
@admin_required
def delete_purchase_order(order_id):
    db = get_db()
    db.execute("DELETE FROM purchase_orders WHERE id = ?", (order_id,))
    db.commit()
    flash('Purchase order deleted.', 'success')
    return redirect(url_for('purchase_orders'))


# --- Warehouse Management (admin only) ---

@app.route('/warehouses')
@admin_required
def warehouses():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT w.*, COUNT(p.id) AS product_count, COALESCE(SUM(p.quantity), 0) AS total_qty
        FROM warehouses w
        LEFT JOIN products p ON p.warehouse_id = w.id
        GROUP BY w.id
        ORDER BY w.name
    """)
    warehouse_list = cursor.fetchall()
    return render_template('warehouses.html', warehouses=warehouse_list)


@app.route('/warehouses/add', methods=['GET', 'POST'])
@admin_required
def add_warehouse():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        location = request.form.get('location', '').strip()
        manager_name = request.form.get('manager_name', '').strip()

        if not name or not location or not manager_name:
            flash('All fields are required.', 'danger')
            return render_template('add_warehouse.html')

        db = get_db()
        db.execute("""
            INSERT INTO warehouses (name, location, manager_name)
            VALUES (?, ?, ?)
        """, (name, location, manager_name))
        db.commit()
        flash('Warehouse added successfully!', 'success')
        return redirect(url_for('warehouses'))

    return render_template('add_warehouse.html')


@app.route('/warehouses/edit/<int:warehouse_id>', methods=['GET', 'POST'])
@admin_required
def edit_warehouse(warehouse_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM warehouses WHERE id = ?", (warehouse_id,))
    warehouse = cursor.fetchone()

    if not warehouse:
        flash('Warehouse not found.', 'danger')
        return redirect(url_for('warehouses'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        location = request.form.get('location', '').strip()
        manager_name = request.form.get('manager_name', '').strip()

        if not name or not location or not manager_name:
            flash('All fields are required.', 'danger')
            return render_template('edit_warehouse.html', warehouse=warehouse)

        db.execute("""
            UPDATE warehouses
            SET name = ?, location = ?, manager_name = ?
            WHERE id = ?
        """, (name, location, manager_name, warehouse_id))
        db.commit()
        flash('Warehouse updated successfully!', 'success')
        return redirect(url_for('warehouses'))

    return render_template('edit_warehouse.html', warehouse=warehouse)


@app.route('/warehouses/delete/<int:warehouse_id>', methods=['POST'])
@admin_required
def delete_warehouse(warehouse_id):
    db = get_db()
    cursor = db.cursor()

    # Prevent deleting the default/only warehouse or one with products
    cursor.execute("SELECT COUNT(*) FROM warehouses")
    if cursor.fetchone()[0] <= 1:
        flash('Cannot delete the last remaining warehouse.', 'danger')
        return redirect(url_for('warehouses'))

    cursor.execute("SELECT COUNT(*) FROM products WHERE warehouse_id = ?", (warehouse_id,))
    if cursor.fetchone()[0] > 0:
        flash('Cannot delete warehouse with active products. Please reassign the products first.', 'danger')
        return redirect(url_for('warehouses'))

    db.execute("DELETE FROM warehouses WHERE id = ?", (warehouse_id,))
    db.commit()
    flash('Warehouse deleted successfully.', 'success')
    return redirect(url_for('warehouses'))


# --- AI Demand Forecasting (admin only) ---

@app.route('/forecasting')
@admin_required
def forecasting():
    db = get_db()
    cursor = db.cursor()
    
    # Get all products for selection dropdown
    cursor.execute("SELECT id, name, quantity, min_stock_threshold FROM products ORDER BY name")
    products = cursor.fetchall()
    
    if not products:
        return render_template('forecasting.html', products=[], selected_product=None)
        
    # Get selected product ID
    product_id = request.args.get('product_id')
    if product_id:
        product_id = int(product_id)
    else:
        product_id = products[0]['id']
        
    cursor.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    selected_product = cursor.fetchone()
    
    # Load past 90 days of transactions
    import datetime
    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=90)
    
    # Query daily stock out movements
    cursor.execute("""
        SELECT date(created_at) as tx_date, SUM(quantity) as qty
        FROM stock_movements
        WHERE product_id = ? AND movement_type = 'out' AND created_at >= ?
        GROUP BY tx_date
    """, (product_id, start_date.strftime('%Y-%m-%d %H:%M:%S')))
    tx_rows = cursor.fetchall()
    
    tx_map = {row['tx_date']: row['qty'] for row in tx_rows}
    
    # Generate daily time series for the past 90 days
    past_dates = [start_date + datetime.timedelta(days=i) for i in range(91)]
    past_qty = [tx_map.get(d.strftime('%Y-%m-%d'), 0) for d in past_dates]
    
    # AI Forecasting Model using scikit-learn
    from sklearn.linear_model import LinearRegression
    import numpy as np
    
    X = np.array([[i] for i in range(len(past_dates))])
    y = np.array(past_qty)
    
    # Train Linear Regression model
    model = LinearRegression()
    model.fit(X, y)
    
    # Predict for the next 30 days
    X_pred = np.array([[i] for i in range(91, 121)])
    y_pred = model.predict(X_pred)
    y_pred = np.clip(y_pred, 0, None)  # demand cannot be negative
    
    # Calculate predicted 30-day demand
    predicted_demand = int(np.ceil(np.sum(y_pred)))
    
    # Reorder Suggestion
    current_stock = selected_product['quantity']
    suggested_reorder = 0
    reorder_needed = False
    if predicted_demand > current_stock:
        reorder_needed = True
        suggested_reorder = predicted_demand - current_stock
        
    # Prepare charts data (past 30 days + next 30 days)
    chart_past_dates = past_dates[-30:]
    chart_past_qty = past_qty[-30:]
    
    chart_future_dates = [today + datetime.timedelta(days=i) for i in range(1, 31)]
    
    # X values for trendline across the chart window (last 30 days + next 30 days)
    X_chart = np.array([[i] for i in range(61, 121)])
    y_chart_trend = model.predict(X_chart)
    y_chart_trend = np.clip(y_chart_trend, 0, None)
    
    # Format labels & datasets for Chart.js
    labels = [d.strftime('%b %d') for d in chart_past_dates] + [d.strftime('%b %d') for d in chart_future_dates]
    actual_sales = list(chart_past_qty) + [None] * 30
    forecast_sales = [float(val) for val in y_chart_trend]
    
    return render_template('forecasting.html',
                           products=products,
                           selected_product=selected_product,
                           predicted_demand=predicted_demand,
                           reorder_needed=reorder_needed,
                           suggested_reorder=suggested_reorder,
                           labels=labels,
                           actual_sales=actual_sales,
                           forecast_sales=forecast_sales)


# --- Notification Settings (admin only) ---

@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    db = get_db()
    if request.method == 'POST':
        admin_email = request.form.get('admin_email', '').strip()
        admin_phone = request.form.get('admin_phone_number', '').strip()
        twilio_sid = request.form.get('twilio_sid', '').strip()
        twilio_token = request.form.get('twilio_auth_token', '').strip()
        twilio_phone = request.form.get('twilio_phone_number', '').strip()
        
        db.execute("INSERT INTO settings (key, value) VALUES ('admin_email', ?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (admin_email,))
        db.execute("INSERT INTO settings (key, value) VALUES ('admin_phone_number', ?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (admin_phone,))
        db.execute("INSERT INTO settings (key, value) VALUES ('twilio_sid', ?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (twilio_sid,))
        db.execute("INSERT INTO settings (key, value) VALUES ('twilio_auth_token', ?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (twilio_token,))
        db.execute("INSERT INTO settings (key, value) VALUES ('twilio_phone_number', ?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (twilio_phone,))
        db.commit()
        
        flash('Notification settings updated successfully!', 'success')
        return redirect(url_for('settings'))
        
    admin_email = get_setting('admin_email')
    admin_phone = get_setting('admin_phone_number')
    twilio_sid = get_setting('twilio_sid')
    twilio_token = get_setting('twilio_auth_token')
    twilio_phone = get_setting('twilio_phone_number')
    
    return render_template('settings.html',
                           admin_email=admin_email,
                           admin_phone_number=admin_phone,
                           twilio_sid=twilio_sid,
                           twilio_auth_token=twilio_token,
                           twilio_phone_number=twilio_phone)


if __name__ == '__main__':
    app.run(debug=app.config.get('DEBUG', True))
